#!/usr/bin/env python3
"""
Universal Financial Statement Extractor
Works with ANY financial statement PDF - automatically detects structure
Extracts income statements, balance sheets, cash flows with high accuracy
"""

import numpy as np
import pytesseract
from pdf2image import convert_from_path
import pandas as pd
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from typing import List, Dict, Tuple, Optional
from PIL import ImageOps

class UniversalFinancialExtractor:
    """Extract financial statements from any PDF automatically"""
    
    def __init__(self, pdf_path: str):
        self.pdf_path = pdf_path
        self.company_name = None
        self.statement_type = None
        self.currency = None
        self.periods = []
        
    def detect_financial_pages(self) -> List[int]:
        """Detect which pages contain financial statements"""
        
        print("\n[1/7] Detecting financial statement pages...")
        
        # Keywords for income statement specifically (prioritize this)
        income_keywords = [
            'revenue from operations', 'other income', 'total income',
            'cost of materials', 'employee benefits expense', 'finance costs',
            'depreciation', 'profit before tax', 'tax expense', 'net profit',
            'income statement', 'statement of profit and loss'
        ]
        
        # Other financial keywords
        other_keywords = [
            'balance sheet', 'assets', 'liabilities', 'equity',
            'cash flow', 'operating activities'
        ]
        
        financial_pages = []
        page_scores = {}
        
        # Quick scan at low DPI to find financial pages
        try:
            pages = convert_from_path(self.pdf_path, dpi=150, first_page=1, last_page=20)
        except:
            pages = convert_from_path(self.pdf_path, dpi=150)
        
        for page_num, page in enumerate(pages, 1):
            # Extract text
            text = pytesseract.image_to_string(page, config='--psm 6').lower()
            
            # Score page based on keywords
            income_score = sum(5 for keyword in income_keywords if keyword in text)  # 5 points each
            other_score = sum(1 for keyword in other_keywords if keyword in text)   # 1 point each
            
            # Check for table-like structure (lots of numbers)
            numbers = re.findall(r'[\d,]+\.[\d]{2}', text)
            number_score = min(len(numbers) // 20, 5)  # Max 5 points
            
            # CRITICAL: Check if this looks like actual income statement (has key line items in sequence)
            has_revenue = 'revenue from operations' in text
            has_income = 'other income' in text or 'total income' in text
            has_expenses = 'cost of materials' in text or 'employee benefits' in text
            has_profit = 'profit before tax' in text or 'net profit' in text
            
            # ANTI-PATTERN: Penalize pages that are just notes/ratios
            is_notes_page = ('notes' in text[:500] or 'additional information' in text[:500]) and text.count('refer note') > 2
            
            # Check if page starts with "notes:" or similar (indicator it's a notes page)
            is_notes_start = text.strip()[:20].startswith('notes')
            
            # Bonus for having the right structure
            structure_bonus = 0
            if has_revenue and has_income and has_expenses and has_profit:
                structure_bonus = 50  # Huge bonus for complete income statement structure
            elif has_revenue and has_expenses and has_profit:
                structure_bonus = 30
            
            # Additional penalty for pages that start with notes
            if is_notes_start:
                structure_bonus -= 30
            
            # Penalty for notes pages
            notes_penalty = -40 if is_notes_page else 0
            
            total_score = income_score + other_score + number_score + structure_bonus + notes_penalty
            
            # Must have at least 3 income keywords OR structure bonus > 0
            if income_score >= 15 or structure_bonus > 0:
                page_scores[page_num] = {
                    'total': total_score,
                    'income': income_score,
                    'numbers': len(numbers),
                    'structure': structure_bonus,
                    'is_notes': is_notes_page
                }
                financial_pages.append(page_num)
                page_type = "Notes/Ratios" if is_notes_page else "Income statement"
                print(f"  ✓ Page {page_num}: {page_type} (score: {total_score}, income keywords: {income_score//5}, structure: {structure_bonus > 0})")
        
        if not financial_pages:
            print("  ! No obvious income statement pages found, scanning all pages...")
            financial_pages = list(range(1, min(len(pages) + 1, 10)))
        else:
            # Sort by score (highest first) and prioritize non-notes pages
            financial_pages = sorted(financial_pages, key=lambda p: (
                not page_scores.get(p, {}).get('is_notes', False),  # Non-notes pages first
                page_scores.get(p, {}).get('total', 0)  # Then by score
            ), reverse=True)
        
        return financial_pages
    
    def extract_metadata(self, text: str):
        """Extract company name, currency, statement type from text"""
        
        lines = text.split('\n')
        
        # Extract company name (usually in first 10 lines)
        if not self.company_name:
            for line in lines[:10]:
                line = line.strip()
                if len(line) > 10 and len(line) < 100:
                    # Look for company indicators
                    if any(word in line.lower() for word in ['limited', 'ltd', 'inc', 'corporation', 'corp']):
                        self.company_name = line
                        break
        
        # Detect currency
        if not self.currency:
            currency_patterns = [
                (r'\(₹\s*in\s+crores?\)', '₹ in crores'),
                (r'\(₹\s*in\s+lakhs?\)', '₹ in lakhs'),
                (r'\(₹\s*in\s+millions?\)', '₹ in millions'),
                (r'\(\$\s*in\s+millions?\)', '$ in millions'),
                (r'\(\$\s*in\s+thousands?\)', '$ in thousands'),
                (r'Rs\.?\s*in\s+crores?', 'Rs. in crores'),
                (r'INR\s+in\s+crores?', 'INR in crores'),
            ]
            
            for pattern, label in currency_patterns:
                if re.search(pattern, text, re.IGNORECASE):
                    self.currency = label
                    break
        
        # Detect statement type
        if not self.statement_type:
            text_lower = text.lower()
            if any(word in text_lower for word in ['income statement', 'profit and loss', 'p&l', 'statement of income']):
                self.statement_type = 'Income Statement'
            elif 'balance sheet' in text_lower:
                self.statement_type = 'Balance Sheet'
            elif 'cash flow' in text_lower:
                self.statement_type = 'Cash Flow Statement'
    
    def detect_column_headers(self, text: str) -> List[str]:
        """Automatically detect time period column headers"""
        
        lines = text.split('\n')
        headers = []
        
        # Patterns for date/period headers
        patterns = [
            r'FY\s*[\'"]?(\d{2})',  # FY 25, FY'25
            r'FY\s*20(\d{2})',  # FY 2025
            r'20(\d{2})-20(\d{2})',  # 2024-2025
            r'(\d{1,2})[/-](\d{1,2})[/-]20(\d{2})',  # 31/12/2025
            r'(Q[1-4])\s+FY\s*[\'"]?(\d{2})',  # Q3 FY26
            r'(Q[1-4])\s+20(\d{2})',  # Q3 2025
            r'(\d+)\s*months?\s+ended',  # 9 months ended
            r'Year\s+ended.*?20(\d{2})',  # Year ended 2025
        ]
        
        found_headers = []
        
        for line in lines[:30]:  # Check first 30 lines
            for pattern in patterns:
                matches = re.findall(pattern, line, re.IGNORECASE)
                if matches:
                    # Extract the matching header text
                    for match in matches:
                        header_text = line[max(0, line.find(str(match[0] if isinstance(match, tuple) else match)) - 20):
                                          line.find(str(match[0] if isinstance(match, tuple) else match)) + 30].strip()
                        if header_text and len(header_text) < 50:
                            found_headers.append(header_text)
        
        # Remove duplicates while preserving order
        seen = set()
        headers = []
        for h in found_headers:
            h_clean = h.strip()
            if h_clean and h_clean not in seen and len(h_clean) > 2:
                seen.add(h_clean)
                headers.append(h_clean)
        
        return headers[:10]  # Max 10 periods
    
    def extract_table_with_structure(self, page_image, page_num: int) -> Optional[pd.DataFrame]:
        """Extract table structure using OCR with position data"""
        
        print(f"  Processing page {page_num}...")
        
        # Convert to grayscale for better OCR
        gray = ImageOps.grayscale(page_image)
        
        # Get OCR data with positions
        ocr_data = pytesseract.image_to_data(gray, output_type=pytesseract.Output.DICT, config='--psm 6')
        
        # Extract text with positions
        text_boxes = []
        for i in range(len(ocr_data['text'])):
            text = ocr_data['text'][i].strip()
            conf = int(ocr_data['conf'][i])
            
            # Increased confidence threshold and clean OCR noise
            if text and conf > 40:  # Confidence > 40% (was 20%)
                # Clean common OCR noise patterns
                cleaned_text = self._clean_ocr_noise(text)
                if cleaned_text:  # Only add if text remains after cleaning
                    text_boxes.append({
                        'text': cleaned_text,
                        'x': ocr_data['left'][i],
                        'y': ocr_data['top'][i],
                        'w': ocr_data['width'][i],
                        'h': ocr_data['height'][i],
                        'conf': conf
                    })
        
        if not text_boxes:
            return None
        
        # Group into rows based on Y position
        rows = self._group_into_rows(text_boxes)
        
        # Group into columns based on X position
        table_data = self._group_into_columns(rows)
        
        if not table_data or len(table_data) < 3:
            return None
        
        # Convert to DataFrame
        df = pd.DataFrame(table_data)
        
        return df
    
    def _clean_ocr_noise(self, text: str) -> str:
        """Clean common OCR noise and artifacts"""
        import re
        
        if not text:
            return ""
        
        # Remove standalone noise characters (common OCR errors)
        noise_patterns = [
            r'^[|_\[\]{}]+$',  # Lines like |, _, [, ], {, } alone
            r'^[|_]+',  # Leading noise like |_, |__, etc.
            r'[|_]+$',  # Trailing noise
        ]
        
        for pattern in noise_patterns:
            text = re.sub(pattern, '', text)
        
        text = text.strip()
        
        # If text is just noise characters, return empty
        if all(c in '|_[]{}' for c in text):
            return ""
        
        # If text is very short and contains mostly noise, skip
        if len(text) <= 2 and any(c in '|_[]{}' for c in text):
            return ""
        
        return text
    
    def _group_into_rows(self, text_boxes: List[Dict]) -> List[List[Dict]]:
        """Group text boxes into rows"""
        
        if not text_boxes:
            return []
        
        # Sort by Y position
        sorted_boxes = sorted(text_boxes, key=lambda x: x['y'])
        
        rows = []
        current_row = [sorted_boxes[0]]
        current_y = sorted_boxes[0]['y']
        
        y_tolerance = 20  # pixels
        
        for box in sorted_boxes[1:]:
            if abs(box['y'] - current_y) < y_tolerance:
                current_row.append(box)
            else:
                rows.append(sorted(current_row, key=lambda x: x['x']))
                current_row = [box]
                current_y = box['y']
        
        if current_row:
            rows.append(sorted(current_row, key=lambda x: x['x']))
        
        return rows
    
    def _group_into_columns(self, rows: List[List[Dict]]) -> List[List[str]]:
        """Group row elements into columns"""
        
        if not rows:
            return []
        
        # Find column boundaries
        all_x_positions = []
        for row in rows:
            for box in row:
                all_x_positions.append(box['x'])
        
        if not all_x_positions:
            return []
        
        all_x_positions = sorted(set(all_x_positions))
        
        # SMART APPROACH: Detect data columns by finding rows with multiple numbers
        # Financial statements have label column on left, then data columns on right
        
        # Strategy: Collect all X positions of numbers, then cluster them
        number_x_positions = []
        
        for row in rows:
            for box in row:
                text = box['text'].strip()
                # Check if this looks like a number (thousands separator, decimals, negatives)
                is_number = False
                clean = text.replace(',', '').replace('.', '').replace('-', '').replace('(', '').replace(')', '')
                if clean.isdigit() and len(clean) >= 2:
                    is_number = True
                
                if is_number:
                    number_x_positions.append(box['x'])
        
        # Get the most common number positions (these are data columns)
        if number_x_positions and len(number_x_positions) >= 5:
            # Find where data columns actually are by clustering number positions
            # The main data columns will have the most numbers concentrated
            number_clusters = self._cluster_positions(sorted(set(number_x_positions)), threshold=150)
            
            # Count how many numbers fall into each cluster
            cluster_counts = []
            for cluster in number_clusters:
                count = sum(1 for x in number_x_positions if abs(x - cluster) < 150)
                cluster_counts.append((cluster, count))
            
            # Sort by count (descending)
            cluster_counts.sort(key=lambda x: x[1], reverse=True)
            
            # Take top clusters (those with significant counts)
            max_count = cluster_counts[0][1] if cluster_counts else 0
            min_threshold = max(5, max_count // 4)  # At least 25% of max count
            data_clusters = [x for x, count in cluster_counts if count >= min_threshold]
            data_clusters.sort()
            
            if len(data_clusters) >= 2:
                # The leftmost data cluster marks the start of data columns
                min_data_x = min(data_clusters)
                
                # Find max label X (everything before first data column)
                label_x_positions = [x for x in all_x_positions if x < min_data_x - 100]
                
                if label_x_positions:
                    max_label_x = max(label_x_positions)
                    label_threshold = (max_label_x + min_data_x) // 2
                    
                    # Create column boundaries:
                    # 1. One column for all labels
                    label_positions = [x for x in all_x_positions if x < label_threshold]
                    column_boundaries = [int(np.mean(label_positions))] if label_positions else []
                    
                    # 2. Data columns - use the identified clusters directly
                    column_boundaries.extend(data_clusters)
                else:
                    # No clear label section, use data clusters only
                    column_boundaries = data_clusters
            else:
                # Not enough distinct data clusters, fall back
                column_boundaries = self._cluster_positions(all_x_positions, threshold=400)
        else:
            # Not enough numbers or no numbers found, fall back to clustering
            column_boundaries = self._cluster_positions(all_x_positions, threshold=400)
        
        # Limit to reasonable number of columns
        if len(column_boundaries) > 12:
            # Too many columns - use larger threshold
            column_boundaries = self._cluster_positions(all_x_positions, threshold=600)
        
        # Assign texts to columns
        table_data = []
        for row in rows:
            row_data = [''] * len(column_boundaries)
            
            for box in row:
                text = box['text']
                
                # Check if this is a number
                is_number = False
                clean = text.replace(',', '').replace('.', '').replace('-', '').replace('(', '').replace(')', '')
                if clean.replace('.', '').isdigit() and len(clean) >= 2:
                    is_number = True
                
                # Find closest column
                distances = [abs(box['x'] - col_x) for col_x in column_boundaries]
                col_idx = distances.index(min(distances))
                
                # IMPORTANT: If this is a number and it's trying to go to first column,
                # but there are other columns, assign it to the closest numeric column instead
                if is_number and col_idx == 0 and len(column_boundaries) > 1:
                    # Find the closest data column (not the first one)
                    data_distances = [abs(box['x'] - col_x) for col_x in column_boundaries[1:]]
                    if data_distances:
                        min_data_dist = min(data_distances)
                        # If the number is closer to a data column than to the label column
                        if min_data_dist < distances[0] + 100:  # 100px tolerance
                            col_idx = data_distances.index(min_data_dist) + 1
                
                # Add text to column
                if row_data[col_idx]:
                    row_data[col_idx] += ' ' + text
                else:
                    row_data[col_idx] = text
            
            # Only add rows with significant content
            if any(len(str(cell).strip()) > 0 for cell in row_data):
                table_data.append(row_data)
        
        return table_data
    
    def _cluster_positions(self, positions: List[int], threshold: int = 50) -> List[int]:
        """Cluster positions to find column boundaries"""
        
        if not positions:
            return []
        
        clusters = []
        current_cluster = [positions[0]]
        
        for pos in positions[1:]:
            # Check distance from FIRST position in cluster, not last
            if pos - current_cluster[0] < threshold:
                current_cluster.append(pos)
            else:
                # Save cluster center and start new cluster
                clusters.append(int(np.mean(current_cluster)))
                current_cluster = [pos]
        
        if current_cluster:
            clusters.append(int(np.mean(current_cluster)))
        
        return clusters
    
    def clean_and_parse_table(self, df: pd.DataFrame, page_text: str) -> pd.DataFrame:
        """Clean and parse the extracted table"""
        
        if df.empty:
            return df
        
        # Remove completely empty rows
        df = df[df.astype(str).apply(lambda x: x.str.strip().str.len().sum(), axis=1) > 0].reset_index(drop=True)
        
        if df.empty:
            return df
        
        # Find header row (contains period information or "Particulars")
        header_idx = 0
        for i, row in df.iterrows():
            first_cell = str(row[0]).lower()
            if any(word in first_cell for word in ['particular', 'description', 'item', 'sr', 's.no']):
                header_idx = i
                break
            # Or if row contains multiple year/period indicators
            row_str = ' '.join(str(cell) for cell in row).lower()
            if row_str.count('fy') >= 2 or row_str.count('20') >= 2 or row_str.count('q') >= 2:
                header_idx = i
                break
        
        # Set headers
        if header_idx > 0:
            df = df.iloc[header_idx:].reset_index(drop=True)
        
        headers = df.iloc[0].tolist()
        
        # Handle duplicate column names by adding suffixes
        seen = {}
        unique_headers = []
        for header in headers:
            if header in seen:
                seen[header] += 1
                unique_headers.append(f"{header}_{seen[header]}")
            else:
                seen[header] = 0
                unique_headers.append(header)
        
        df.columns = unique_headers
        df = df.iloc[1:].reset_index(drop=True)
        
        # Clean numeric columns (detect and convert)
        # Use positional indexing to avoid issues with duplicate column names
        for col_idx in range(1, len(df.columns)):
            try:
                df.iloc[:, col_idx] = df.iloc[:, col_idx].apply(self._clean_numeric_value)
            except:
                pass  # Skip columns that can't be cleaned
        
        # Extract metadata from page text
        self.extract_metadata(page_text)
        
        return df
    
    def _clean_numeric_value(self, value):
        """Clean and convert numeric values"""
        
        # Handle pandas Series (shouldn't happen but just in case)
        if isinstance(value, pd.Series):
            return value
        
        if pd.isna(value) or value == '' or value is None:
            return None
        
        value = str(value).strip()
        
        # Handle empty or very short values
        if len(value) < 1:
            return None
        
        # Handle negative numbers in parentheses
        if value.startswith('(') and value.endswith(')'):
            value = '-' + value[1:-1]
        
        # Remove common formatting
        value = value.replace(',', '')
        value = value.replace('₹', '')
        value = value.replace('$', '')
        value = value.replace('Rs.', '')
        value = value.replace('–', '-')
        value = value.replace('—', '-')
        
        # Remove trailing non-numeric characters (like | ] [ etc.)
        # But preserve negative sign at start
        value = value.strip()
        while value and not value[-1].isdigit() and value[-1] not in '0123456789':
            value = value[:-1].strip()
        
        # Remove leading non-numeric characters except negative sign
        while value and value[0] not in '-0123456789' and not value[0].isdigit():
            value = value[1:].strip()
        
        # Try to convert to float
        try:
            return float(value)
        except:
            return value
    
    def extract(self, output_path: Optional[str] = None) -> pd.DataFrame:
        """Main extraction method - works with ANY financial PDF"""
        
        print("="*80)
        print("UNIVERSAL FINANCIAL STATEMENT EXTRACTOR")
        print("="*80)
        print(f"\nPDF: {self.pdf_path}")
        
        # Step 1: Detect financial pages
        financial_pages = self.detect_financial_pages()
        
        # Step 2: Extract from each page
        print(f"\n[2/7] Extracting tables from {len(financial_pages)} page(s)...")
        
        all_tables = []
        
        for page_num in financial_pages:
            # Convert page to image
            pages = convert_from_path(
                self.pdf_path,
                dpi=300,
                first_page=page_num,
                last_page=page_num
            )
            
            if not pages:
                print(f"    ✗ Failed to convert page {page_num}")
                continue
            
            page_image = pages[0]
            
            # Extract text for metadata
            page_text = pytesseract.image_to_string(page_image, config='--psm 6')
            
            # Extract table
            df = self.extract_table_with_structure(page_image, page_num)
            
            if df is None:
                print(f"    ✗ No table structure detected on page {page_num}")
                continue
            
            if df.empty:
                print(f"    ✗ Empty table on page {page_num}")
                continue
            
            # Clean and parse
            df = self.clean_and_parse_table(df, page_text)
            
            if df.empty:
                print(f"    ✗ Table became empty after cleaning on page {page_num}")
            elif len(df.columns) < 2:
                print(f"    ✗ Insufficient columns ({len(df.columns)}) on page {page_num}")
            else:
                all_tables.append((page_num, df))
                print(f"    ✓ Extracted {len(df)} rows x {len(df.columns)} columns")
        
        if not all_tables:
            print("\n❌ No financial tables extracted")
            return pd.DataFrame()
        
        # Step 3: Combine tables
        print(f"\n[3/7] Combining {len(all_tables)} table(s)...")
        
        # Use the table from the highest-scoring page (not just largest)
        # Priority: page with best score from financial_pages order
        primary_table = all_tables[0]  # First in list has highest score
        combined_df = primary_table[1]
        
        print(f"  ✓ Primary table: {len(combined_df)} rows x {len(combined_df.columns)} cols from page {primary_table[0]}")
        
        # Step 4: Post-processing
        print(f"\n[4/7] Post-processing data...")
        combined_df = self._post_process(combined_df)
        
        # Step 5: Create Excel
        if output_path:
            print(f"\n[5/7] Creating Excel file...")
            self.create_excel(combined_df, output_path)
        
        # Step 6: Summary
        print(f"\n[6/7] Extraction summary...")
        print(f"  Company: {self.company_name or 'Not detected'}")
        print(f"  Statement Type: {self.statement_type or 'Not detected'}")
        print(f"  Currency: {self.currency or 'Not detected'}")
        print(f"  Rows: {len(combined_df)}")
        print(f"  Columns: {len(combined_df.columns)}")
        
        print(f"\n[7/7] ✅ EXTRACTION COMPLETE")
        
        return combined_df
    
    def _post_process(self, df: pd.DataFrame) -> pd.DataFrame:
        """Post-process the dataframe"""
        
        if df.empty:
            return df
        
        # Clean OCR noise from the first column (Particulars)
        df.iloc[:, 0] = df.iloc[:, 0].apply(self._clean_particulars_text)
        
        # Remove rows where first column is empty or too short
        df = df[df.iloc[:, 0].astype(str).str.len() > 2].reset_index(drop=True)
        
        # Remove duplicate rows
        df = df.drop_duplicates().reset_index(drop=True)
        
        return df
    
    def _clean_particulars_text(self, text):
        """Clean OCR noise from particulars column"""
        import re
        
        if pd.isna(text) or text == '':
            return text
        
        text = str(text)
        
        # Extract text from noisy brackets
        # e.g., "[| __|Totalincome 699.29]" -> "Totalincome 699.29"
        bracket_match = re.search(r'\[([^\]]*)\]', text)
        if bracket_match:
            bracket_content = bracket_match.group(1)
            # If bracket contains noise characters, extract the clean part
            if any(c in bracket_content for c in '|_'):
                # Remove the noise and keep alphabetic content
                clean_content = re.sub(r'[|_]+\s*', '', bracket_content)
                text = clean_content
        
        # Remove common OCR artifacts at beginning/end
        text = re.sub(r'^[|_\[\]{}]+\s*', '', text)  # Leading noise
        text = re.sub(r'\s*[|_\[\]{}]+$', '', text)  # Trailing noise
        
        # Remove numbers from the particulars column (they belong in value columns)
        # But keep line item numbers like "1 Income", "2 Expenses"
        # Remove trailing numbers and decimals (but not if they're part of the description)
        text = re.sub(r'\s+[\d,\.]+\s*$', '', text)
        
        # Remove trailing punctuation that's likely OCR noise
        text = re.sub(r'[,;:]+$', '', text)
        
        # Clean up multiple spaces
        text = re.sub(r'\s+', ' ', text)
        
        return text.strip()
    
    def create_excel(self, df: pd.DataFrame, output_path: str):
        """Create professional Excel file"""
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financial Statement"
        
        # Title
        ws['A1'] = self.company_name or "Financial Statement"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center")
        num_cols = len(df.columns)
        ws.merge_cells(f'A1:{chr(64+num_cols)}1')
        
        ws['A2'] = self.statement_type or "Financial Results"
        ws['A2'].font = Font(size=12, italic=True)
        ws['A2'].alignment = Alignment(horizontal="center")
        ws.merge_cells(f'A2:{chr(64+num_cols)}2')
        
        ws['A3'] = self.currency or "(All amounts)"
        ws['A3'].font = Font(size=10, italic=True, color="666666")
        ws['A3'].alignment = Alignment(horizontal="center")
        ws.merge_cells(f'A3:{chr(64+num_cols)}3')
        
        # Headers
        start_row = 5
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value = col_name
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Data
        for row_idx, row_data in enumerate(df.itertuples(index=False), start_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                if col_idx == 1:
                    cell.value = value
                    cell.alignment = Alignment(horizontal="left", wrap_text=True)
                else:
                    if value is None or value == '':
                        cell.value = None
                    elif isinstance(value, (int, float)):
                        cell.value = float(value)
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal="right")
                    else:
                        cell.value = value
                        cell.alignment = Alignment(horizontal="right")
        
        # Column widths
        ws.column_dimensions['A'].width = 60
        for col_idx in range(2, num_cols + 1):
            ws.column_dimensions[chr(64 + col_idx)].width = 15
        
        # Borders
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        for row in ws[f'A{start_row}:{chr(64+num_cols)}{start_row + len(df)}']:
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border
        
        # Save
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        print(f"  ✓ Excel saved: {output_path}")

def main():
    """Main function for testing"""
    
    import sys
    
    if len(sys.argv) > 1:
        pdf_path = sys.argv[1]
    else:
        # Default test files
        test_files = [
            "Dabur Quaterly Financial Statements.pdf",
            "Tata Motors Quarterly Financial Statements.pdf"
        ]
        
        # Use first available
        pdf_path = None
        for f in test_files:
            if Path(f).exists():
                pdf_path = f
                break
        
        if not pdf_path:
            print("Usage: python universal_financial_extractor.py <pdf_file>")
            print("\nOr place one of these files in the current directory:")
            for f in test_files:
                print(f"  - {f}")
            return
    
    # Extract
    extractor = UniversalFinancialExtractor(pdf_path)
    
    # Generate output filename
    pdf_name = Path(pdf_path).stem
    output_path = f"output/{pdf_name}_universal_extraction.xlsx"
    
    # Run extraction
    df = extractor.extract(output_path)
    
    # Display
    if not df.empty:
        print("\n" + "="*80)
        print("EXTRACTED DATA PREVIEW")
        print("="*80)
        print(df.to_string(index=False, max_rows=30))
        print("\n" + "="*80)
    
    return df

if __name__ == "__main__":
    main()
